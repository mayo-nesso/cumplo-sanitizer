import csv
from datetime import datetime
from typing import Optional

import ipywidgets as widgets
import openpyxl
import pandas as pd
from IPython.display import display
from pyxirr import InvalidPaymentsError, xirr


def find_negative_earning_ids(movs: pd.DataFrame) -> list[str]:
    # Filter records where total_earnings is negative
    negative_investments = movs.groupby("RemateID").apply(
        lambda group: group["Abono"].sum() - group["Cargo"].sum() < 0
    )
    return negative_investments[negative_investments.values].index.tolist()


# Get rates from a collection/df of movements...
def get_rates(
    flows: pd.DataFrame,
) -> (Optional[int], Optional[float], Optional[float], Optional[float]):
    diff_days = 0
    mrate_iir = 0
    rate_iir_yr = 0
    rate_xir = 0
    sorted_values = flows.sort_values(by="Fecha", ascending=True)

    fechas = sorted_values["Fecha"]
    if len(fechas) == 0:
        return (None, None, None, None)

    diff_days = (max(fechas) - min(fechas)).days

    mrate_iir = sorted_values["Abono"].sum() / sorted_values["Cargo"].sum() - 1
    if mrate_iir == 0:
        return (diff_days, mrate_iir, None, None)

    lowest_possible = -1.0
    rate_iir_yr = max(lowest_possible, mrate_iir * 360 / (1 if diff_days <= 1 else (diff_days - 1)))
    try:
        flow_amounts = -1 * sorted_values["Cargo"] + sorted_values["Abono"]
        rate_xir = xirr(fechas, flow_amounts)
    except InvalidPaymentsError:
        rate_xir = None

    return (diff_days, mrate_iir, rate_iir_yr, rate_xir)


# fix missing elements discovered by browsing records...
def insert_fix(original_df: pd.DataFrame, fixdata_csv_path: str) -> pd.DataFrame:
    def create_return_row(r_id: str, actor: str, date: str, abono: str, cargo: str) -> dict:
        row = {
            "Fecha": pd.Timestamp(date),
            "Cargo": int(cargo),
            "Abono": int(abono),
            "Descripción": f"fix_ Pago de inversión, solicitud Credito {actor} {r_id}",
            "Tipo": "Fix",
            "Solicitud": f"Credito {actor} {r_id}",
            "RemateID": r_id,
            "Actor": actor,
        }
        return row

    def get_fix_data(fixdata_csv_path: str) -> list[str]:
        fix_data = []
        with open(fixdata_csv_path, newline="") as csv_file:
            csv_reader = csv.reader(csv_file)

            # read header...
            _ = next(csv_reader, None)

            # Read all the fixes and store them on fix_data
            for row in csv_reader:
                fix_data.append(row)
        return fix_data

    if fixdata_csv_path is None:
        print("No fixdata csv path specified. No changes made.")
        return original_df

    fix_data = get_fix_data(fixdata_csv_path)

    new_rows = []
    for data in fix_data:
        row = create_return_row(data[0], data[1], data[2], data[3], data[4])
        new_rows.append(row)

    fixes_df = pd.DataFrame(new_rows)
    new_df = pd.concat([original_df, fixes_df])
    new_df = new_df.reset_index(drop=True)
    return new_df


def explore_by_id(df: pd.DataFrame, filter_by_ids: list[str] = None):
    if filter_by_ids is not None:
        mask = df["RemateID"].isin(filter_by_ids)
        df = df[mask]

    dfg = df.groupby("RemateID")

    r_ids = list(dfg.groups.keys())
    prev_button = widgets.Button(description="Prev")
    next_button = widgets.Button(description="Next")

    id_slider = widgets.IntSlider(
        value=-1,
        min=0,
        max=len(r_ids) - 1,
        step=1,
        description="R_ID Index:",
        disabled=False,
        continuous_update=False,
        orientation="horizontal",
        readout=True,
        readout_format="d",
    )

    labl_id = widgets.Label()
    labl_stats = widgets.Label()
    labl_amounts = widgets.Label()

    index_text = widgets.IntText()
    index_text.layout.display = "none"
    index_text.value = 0

    def on_next_button_clicked(_):
        index_text.value = (index_text.value + 1) % len(r_ids)

    def on_prev_button_clicked(_):
        index_text.value = (index_text.value - 1 + len(r_ids)) % len(r_ids)

    widgets.jslink((index_text, "value"), (id_slider, "value"))
    # linking button and function together using a button's method
    prev_button.on_click(on_prev_button_clicked)
    next_button.on_click(on_next_button_clicked)

    df_wrapper = widgets.Output()

    def _interactive_df(index_text_value):
        # Get investment r_id!
        r_id = r_ids[index_text_value]
        # and display it!
        labl_id.value = f"RemateID: [{r_id}], Index: [{index_text_value}/{len(r_ids)-1}]"

        # Get selected investment r_id!
        selected_group = dfg.get_group(r_id)
        sorted_values = selected_group.sort_values(by="Fecha", ascending=True)
        # and display them!
        df_wrapper.clear_output()
        df_wrapper.append_display_data(sorted_values)

        # Get Earnings, Charges and diff amounts!
        tot_earnings = sorted_values["Abono"].sum()
        tot_charges = sorted_values["Cargo"].sum()
        diff = tot_earnings - tot_charges
        # and display them!
        labl_amounts.value = (
            f"Earnings [{tot_earnings:,}] Charges [{tot_charges:,}]; delta: [{diff:,}]"
        )

        # Get rates!
        diff_days, mtasa_iir, tasa_iir_yr, tasa_xir = get_rates(selected_group)
        # and display them!
        result_label = f"Days:[{diff_days}]"
        result_label += f" - Rate : [{mtasa_iir * 100:.2f}%]" if mtasa_iir is not None else ""
        result_label += f", Yr [{tasa_iir_yr * 100:.2f}%]" if tasa_iir_yr else ""
        result_label += f"- TIR: [{tasa_xir * 100:.2f}%]" if tasa_xir else ""
        labl_stats.value = result_label

    display(
        widgets.VBox(
            [
                widgets.HBox([prev_button, next_button, id_slider, labl_id]),
                labl_amounts,
                labl_stats,
                df_wrapper,
                index_text,
            ]
        )
    )

    widgets.interact(_interactive_df, index_text_value=index_text)


# Analize flows file line by line, and extract: active, late, and uncollectible ids
# Active means at least one flow in gray (text color = gray)
# Late  means at least one flow in red (text color = red)
# uncollectible are all those investments that have a pending flow older than 'grace_period_days'
def extract_active_and_late_ids(
    flows_file_path: str, grace_period_days
) -> (list[str], list[str], list[str], list[str]):
    # Colors and meaning !!
    c_red = "FFCE494F"  # red | pending!!
    c_gray = "FF808080"  # gray | expected, future payment
    # c_green = 'FF95BB65' # green | payment on time!!
    # c_orange = 'FFFFA500' # orange | payment late, but payed :)

    workbook = openpyxl.load_workbook(flows_file_path, data_only=True)
    sheet = workbook.active

    # Store all ids on document.
    # Some investment that are recently payed could not be registered here,
    # so we will have to know what ids exist in the dococument
    all_ids = set()

    # Get those that are currentlty active and on track
    # ie, it has 'future payments' => *grey* cells!
    active_ids = set()

    # Also, we know that *red* cells means a positive flow that never happend
    late_ids = set()

    # And for those that are late, we will compare the late flow with the current date and
    # see if the diference is more than `grace_period_days`.
    # If its more than that, we declare the investment as uncollectible :(
    uncollectible_ids = set()

    # row range starts on '2' to skip headers!
    sheet_headers_lines = 1
    sheet_footer_lines = 5
    # On Pandas; when possible; iterate over each column and then check each row.
    # This provides a better performance (here ~5x aprox!).
    for column in range(1, sheet.max_column + 1):
        for row in range(1 + sheet_headers_lines, sheet.max_row + 1 - sheet_footer_lines):
            cell = sheet.cell(row=row, column=column)
            if cell.value is None or cell.value == "":
                continue

            font_color = cell.font.color
            if font_color is None:
                continue

            # get row-investment_id
            id = str(sheet.cell(row=row, column=1).value)

            # store this id in the all_ids set...
            all_ids.add(id)

            # Some payment for the future => Currently active
            if font_color.rgb == c_gray:
                active_ids.add(id)
                continue

            # Some late pending payment => (late_and_active and late_and_uncollectible...)
            if font_color.rgb == c_red:
                late_ids.add(id)

                # Check the date!
                date = sheet.cell(row=1, column=column).value
                is_uncollectible = (datetime.now().date() - date.date()).days > grace_period_days

                if is_uncollectible:
                    uncollectible_ids.add(id)

    # Close file!
    workbook.close()
    # return elements as lists
    return (list(all_ids), list(active_ids), list(late_ids), list(uncollectible_ids))


# Extract all those ids where the diff of Abonos and Cargos is less or equal than 'despreciable_amount'
def extract_unexecuted(df: pd.DataFrame, not_present_in_flows_ids: list[str], despreciable_amount: int) -> list[str]:
    dfg = df.groupby("RemateID")

    unexecuted_ids = set()
    for group_key, df_group in dfg:
        earnings = df_group["Abono"].sum()
        cost = df_group["Cargo"].sum()
        investment_diff = earnings - cost

        if abs(investment_diff) <= despreciable_amount or group_key in not_present_in_flows_ids:
            unexecuted_ids.add(group_key)
        elif (
            df_group["Descripción"]
            .str.startswith("Devolución de fondos por crédito no concretado")
            .any()
        ):
            unexecuted_ids.add(group_key)

    return list(unexecuted_ids)



def extract_just_payed(df: pd.DataFrame, not_present_in_flows_ids: list[str], considerable_amount: int) -> list[str]:
    dfg = df.groupby("RemateID")

    just_payed = set()
    for group_key, df_group in dfg:
        earnings = df_group["Abono"].sum()
        cost = df_group["Cargo"].sum()
        investment_diff = earnings - cost

        if investment_diff <= - 1 * abs(considerable_amount) and group_key in not_present_in_flows_ids:
            just_payed.add(group_key)
        

    return list(just_payed)
