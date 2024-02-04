import csv
from typing import Optional

import ipywidgets as widgets
import openpyxl
import pandas as pd

# import some_utils
from IPython.display import display
from pyxirr import InvalidPaymentsError, xirr

# from cumplo_sanitizer.src import some_utils ## works for tests but it doesnt work for jypyter!!
from . import some_utils


def find_negative_earning_ids(movs: pd.DataFrame) -> list[str]:
    """
    Identify and return a list of IDs with negative earnings.

    This function filters through a DataFrame of financial movements and identifies
    those entities (grouped by 'RemateID') whose total 'Abono' minus 'Cargo' results in a
    negative value, indicating negative earnings.

    Parameters
    ----------
    movs : pd.DataFrame
        A DataFrame containing financial movements. It should have at least the
        following columns: 'RemateID', 'Abono', and 'Cargo'.
        - 'RemateID' is an identifier for the entity.
        - 'Abono' represents credits to the account.
        - 'Cargo' represents debits from the account.

    Returns
    -------
    list[str]
        A list of 'RemateID' values corresponding to those entities with negative earnings.

    Examples
    --------
    >>> data = {'RemateID': ['A', 'A', 'B', 'B'],
                'Abono': [100, 200, 300, 400],
                'Cargo': [300, 100, 200, 500]}
    >>> df = pd.DataFrame(data)
    >>> find_negative_earning_ids(df)
    ['B']
    """
    # Filter records where total_earnings is negative
    negative_investments = movs.groupby("RemateID").apply(
        lambda group: group["Abono"].sum() - group["Cargo"].sum() < 0
    )
    return negative_investments[negative_investments.values].index.tolist()


# Get rates from a collection/df of movements...
def _get_rates(
    flows: pd.DataFrame,
) -> (Optional[int], Optional[float], Optional[float], Optional[float]):
    diff_days = 0
    mrate_iir = 0
    rate_iir_yr = 0
    rate_xir = 0
    sorted_values = flows.sort_values(by="Fecha", ascending=True)

    fechas = sorted_values["Fecha"]
    if len(fechas) == 0:
        return (None, None, None, None)

    diff_days = (max(fechas) - min(fechas)).days

    mrate_iir = sorted_values["Abono"].sum() / sorted_values["Cargo"].sum() - 1
    if mrate_iir == 0:
        return (diff_days, mrate_iir, None, None)

    lowest_possible = -1.0
    rate_iir_yr = max(lowest_possible, mrate_iir * 360 / (1 if diff_days <= 1 else (diff_days - 1)))
    try:
        flow_amounts = -1 * sorted_values["Cargo"] + sorted_values["Abono"]
        rate_xir = xirr(fechas, flow_amounts)
    except InvalidPaymentsError:
        rate_xir = None

    return (diff_days, mrate_iir, rate_iir_yr, rate_xir)


def _create_return_row(r_id: str, actor: str, date: str, abono: str, cargo: str) -> dict:
    row = {
        "Fecha": pd.Timestamp(date),
        "Cargo": int(cargo),
        "Abono": int(abono),
        "Descripción": f"fix_ Pago de inversión, solicitud Credito {actor} {r_id}",
        "Tipo": "Fix",
        "Solicitud": f"Credito {actor} {r_id}",
        "RemateID": r_id,
        "Actor": actor,
    }
    return row


def _get_fix_data(fixdata_csv_path: str) -> list[str]:
    fix_data = []
    with open(fixdata_csv_path, newline="") as csv_file:
        csv_reader = csv.reader(csv_file)

        # read header...
        _ = next(csv_reader, None)

        # Read all the fixes and store them on fix_data
        for row in csv_reader:
            fix_data.append(row)
    return fix_data


# fix missing elements discovered by browsing records...
def insert_fix(original_df: pd.DataFrame, fixdata_csv_path: str) -> pd.DataFrame:
    """
    Insert additional rows into a DataFrame from a CSV file.

    This function reads data from a specified CSV file and appends each row as a new
    entry into the original DataFrame. The data from the CSV file is processed by
    the '_create_return_row' function for each row before being appended.

    Parameters
    ----------
    original_df : pd.DataFrame
        The original DataFrame to which the new rows will be appended.

    fixdata_csv_path : str
        The file path to the CSV file containing the data to be inserted.
        If this is None, the function will return the original DataFrame unchanged.

    Returns
    -------
    pd.DataFrame
        A new DataFrame consisting of the original DataFrame with the additional
        rows appended from the CSV file. If no CSV path is provided, returns the
        original DataFrame unmodified.

    Notes
    -----
    - This function relies on '_get_fix_data' to read the CSV file and
      '_create_return_row' to process each row from the CSV file.
    - If 'fixdata_csv_path' is None, the function prints a message and returns
      the original DataFrame without any modifications.

    Examples
    --------
    >>> original_data = {'col1': [1, 2], 'col2': [3, 4]}
    >>> original_df = pd.DataFrame(original_data)
    >>> insert_fix(original_df, 'path/to/fixdata.csv')
    # Returns a DataFrame with rows from 'fixdata.csv' appended to 'original_df'.
    """
    if fixdata_csv_path is None:
        print("No fixdata csv path specified. No changes made.")
        return original_df

    fix_data = _get_fix_data(fixdata_csv_path)

    new_rows = []
    for data in fix_data:
        row = _create_return_row(data[0], data[1], data[2], data[3], data[4])
        new_rows.append(row)

    fixes_df = pd.DataFrame(new_rows)
    fixes_abono = fixes_df["Abono"].sum()
    print(f"Fixing {len(fixes_df)} rows, For a total of ${fixes_abono:,} ")
    new_df = pd.concat([original_df, fixes_df])
    new_df = new_df.reset_index(drop=True)

    return new_df


def explore_by_id(df: pd.DataFrame, filter_by_ids: list[str] = None):
    """
    Interactive exploration of a DataFrame filtered by specific IDs.

    This function allows for the interactive exploration of a given DataFrame,
    with the option to filter the data by a list of IDs. It displays the data
    for each ID along with calculated statistics and allows navigation between
    different IDs using widgets.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame to be explored. It should contain a column 'RemateID'.

    filter_by_ids : list[str], optional
        A list of 'RemateID' values to filter the DataFrame.
        If None (default), no filtering is applied.

    Notes
    -----
    - The function relies on ipywidgets for the interactive components and
      assumes it is being used within an IPython or Jupyter environment.
    - 'RemateID' is used as the key for grouping and filtering the DataFrame.
    - The function utilizes the '_get_rates' function to calculate rates for each group.

    Examples
    --------
    >>> data = {'RemateID': ['ID1', 'ID1', 'ID2'],
                'Fecha': ['2023-01-01', '2023-01-02', '2023-01-03'],
                'Abono': [100, 200, 300],
                'Cargo': [50, 60, 70]}
    >>> df = pd.DataFrame(data)
    >>> explore_by_id(df, filter_by_ids=['ID1', 'ID2'])
    # This will display interactive widgets for exploring IDs 'ID1' and 'ID2'.
    """
    if filter_by_ids is None or len(filter_by_ids) == 0:
        print("filter_by_ids is None or empty... - Aborting...")
        return

    mask = df["RemateID"].isin(filter_by_ids)
    df = df[mask]

    dfg = df.groupby("RemateID")

    r_ids = list(dfg.groups.keys())
    prev_button = widgets.Button(description="Prev")
    next_button = widgets.Button(description="Next")

    id_slider = widgets.IntSlider(
        value=-1,
        min=0,
        max=len(r_ids) - 1,
        step=1,
        description="R_ID Index:",
        disabled=False,
        continuous_update=False,
        orientation="horizontal",
        readout=True,
        readout_format="d",
    )

    labl_id = widgets.Label()
    labl_stats = widgets.Label()
    labl_amounts = widgets.Label()

    index_text = widgets.IntText()
    index_text.layout.display = "none"
    index_text.value = 0

    def on_next_button_clicked(_):
        index_text.value = (index_text.value + 1) % len(r_ids)

    def on_prev_button_clicked(_):
        index_text.value = (index_text.value - 1 + len(r_ids)) % len(r_ids)

    widgets.jslink((index_text, "value"), (id_slider, "value"))
    # linking button and function together using a button's method
    prev_button.on_click(on_prev_button_clicked)
    next_button.on_click(on_next_button_clicked)

    df_wrapper = widgets.Output()

    def _interactive_df(index_text_value):
        # Get investment r_id!
        r_id = r_ids[index_text_value]
        # and display it!
        labl_id.value = f"RemateID: [{r_id}], Index: [{index_text_value}/{len(r_ids)-1}]"

        # Get selected investment r_id!
        selected_group = dfg.get_group(r_id)
        sorted_values = selected_group.sort_values(by="Fecha", ascending=True)
        # and display them!
        df_wrapper.clear_output()
        df_wrapper.append_display_data(sorted_values)

        # Get Earnings, Charges and diff amounts!
        tot_earnings = sorted_values["Abono"].sum()
        tot_charges = sorted_values["Cargo"].sum()
        diff = tot_earnings - tot_charges
        # and display them!
        labl_amounts.value = (
            f"Earnings [{tot_earnings:,}] Charges [{tot_charges:,}]; delta: [{diff:,}]"
        )

        # Get rates!
        diff_days, mtasa_iir, tasa_iir_yr, tasa_xir = _get_rates(selected_group)
        # and display them!
        result_label = f"Days:[{diff_days}]"
        result_label += f" - Rate : [{mtasa_iir * 100:.2f}%]" if mtasa_iir is not None else ""
        result_label += f", Yr [{tasa_iir_yr * 100:.2f}%]" if tasa_iir_yr else ""
        result_label += f"- TIR: [{tasa_xir * 100:.2f}%]" if tasa_xir else ""
        labl_stats.value = result_label

    display(
        widgets.VBox(
            [
                widgets.HBox([prev_button, next_button, id_slider, labl_id]),
                labl_amounts,
                labl_stats,
                df_wrapper,
                index_text,
            ]
        )
    )

    widgets.interact(_interactive_df, index_text_value=index_text)


def extract_active_and_late_ids(
    flows_file_path: str, grace_period_days
) -> (list[str], list[str], list[str], list[str]):
    """
    Analyze a spreadsheet of financial flows and categorize investments based on their status.

    The function reads from an Excel file, identifying investments as 'active', 'late',
    or 'uncollectible' based on their payment status, which is indicated by the text color
    in the spreadsheet. It categorizes investments into these groups and returns lists
    of IDs for each category.

    Parameters
    ----------
    flows_file_path : str
        The path to the Excel file containing the flow data.

    grace_period_days : int
        The number of days to use as the grace period when determining if an investment
        is uncollectible.

    Returns
    -------
    tuple of list[str]
        A tuple containing four lists:
        1. All investment IDs found in the document.
        2. IDs of active investments (payments expected in the future).
        3. IDs of late investments (payments overdue but not yet declared uncollectible).
        4. IDs of uncollectible investments (payments overdue beyond the grace period).

    Notes
    -----
    - This function relies on the 'openpyxl' library to read the Excel file.
    - The function uses the 'is_date_past_grace_period' method from 'some_utils'
      to determine if an investment is uncollectible.
    - Text colors in the spreadsheet are used to determine the status of payments:
      'gray' for active, 'red' for late, and other colors are not considered in this context.

    Examples
    --------
    >>> extract_active_and_late_ids('path/to/flows.xlsx', 30)
    # Returns four lists of investment IDs categorized as all, active, late, and uncollectible.
    """
    # Active means at least one flow in gray (text color = gray)
    # Late  means at least one flow in red (text color = red)
    # uncollectible are all those investments that have a 'late' flow older than 'grace_period_days'

    # Colors and meaning !!
    c_red = "FFCE494F"  # red | pending!!
    c_gray = "FF808080"  # gray | expected, future payment
    # c_green = 'FF95BB65' # green | payment on time!!
    # c_orange = 'FFFFA500' # orange | payment late, but payed :)

    workbook = openpyxl.load_workbook(flows_file_path, data_only=True)
    sheet = workbook.active

    # Store all ids on document.
    # Some investment that are recently payed could not be registered here,
    # so we will have to know what ids exist in the dococument
    all_ids = set()

    # Get those that are currentlty active and on track
    # ie, it has 'future payments' => *grey* cells!
    active_ids = set()

    # Also, we know that *red* cells means a positive flow that never happend
    late_ids = set()

    # And for those that are late, we will compare the late flow with the current date and
    # see if the diference is more than `grace_period_days`.
    # If its more than that, we declare the investment as uncollectible :(
    uncollectible_ids = set()

    # row range starts on '2' to skip headers!
    sheet_headers_lines = 1
    sheet_footer_lines = 5
    # On Pandas; when possible; iterate over each column and then check each row.
    # This provides a better performance (here ~5x aprox!).
    for column in range(1, sheet.max_column + 1):
        for row in range(1 + sheet_headers_lines, sheet.max_row + 1 - sheet_footer_lines):
            cell = sheet.cell(row=row, column=column)
            if cell.value is None or cell.value == "":
                continue

            font_color = cell.font.color
            if font_color is None:
                continue

            # get row-investment_id
            id = str(int(sheet.cell(row=row, column=1).value))

            # store this id in the all_ids set...
            all_ids.add(id)

            # Some payment for the future => Currently active
            if font_color.rgb == c_gray:
                active_ids.add(id)
                continue

            # Some late pending payment => (late_and_active and late_and_uncollectible...)
            if font_color.rgb == c_red:
                late_ids.add(id)

                # Check the date!
                date = sheet.cell(row=1, column=column).value
                is_uncollectible = some_utils.is_date_past_grace_period(grace_period_days, date)

                if is_uncollectible:
                    uncollectible_ids.add(id)

    # Close file!
    workbook.close()
    # return elements as lists
    return (list(all_ids), list(active_ids), list(late_ids), list(uncollectible_ids))


def extract_unexecuted(df: pd.DataFrame, despreciable_amount: int) -> list[str]:
    """
    Extract investment IDs where the net investment (Abonos minus Cargos) is less or
    equal than 'despreciable_amount'

    This function processes a DataFrame of investment data, grouping by 'RemateID'. It identifies
    investments where the difference between total earnings ('Abono') and total cost ('Cargo')
    is less than or equal to a specified 'despreciable_amount'

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing investment data. Expected to have columns 'RemateID',
        'Abono', 'Cargo', and 'Descripción'.

    despreciable_amount : int
        The threshold amount below which the difference between earnings and cost is
        considered negligible.

    Returns
    -------
    list[str]
        List of 'RemateID's where the net investment is less than or equal to the
        despreciable amount

    Examples
    --------
    >>> data = {'RemateID': ['ID1', 'ID2'],
                'Abono': [1000, 50],
                'Cargo': [980, 60],
                'Descripción': ['Investment', 'Devolución de fondos por crédito no concretado']}
    >>> df = pd.DataFrame(data)
    >>> extract_unexecuted(df, 50)
    # Returns a list of IDs, including 'ID2'.
    """
    dfg = df.groupby("RemateID")

    unexecuted_ids = set()
    for group_key, df_group in dfg:
        earnings = df_group["Abono"].sum()
        cost = df_group["Cargo"].sum()
        investment_diff = earnings - cost

        if abs(investment_diff) <= abs(despreciable_amount):
            unexecuted_ids.add(group_key)
        elif (
            df_group["Descripción"]
            .str.startswith("Devolución de fondos por crédito no concretado")
            .any()
        ):
            unexecuted_ids.add(group_key)

    return list(unexecuted_ids)


def extract_just_payed(
    df: pd.DataFrame, not_present_in_flows_ids: list[str], considerable_amount: int
) -> list[str]:
    """
    Extract IDs of investments that have recently been paid off, based on a specified amount.

    This function filters the DataFrame for specific investment IDs and then groups
    it by 'RemateID'. It identifies those investments where the difference between
    total earnings ('Abono') and total cost ('Cargo') is less than or equal to
    the negative of a given 'considerable_amount'. Only investments listed in
    'not_present_in_flows_ids' are considered.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing investment data. Expected to have columns 'RemateID',
        'Abono', and 'Cargo'.

    not_present_in_flows_ids : list[str]
        List of 'RemateID's to be considered for identifying if just paid.

    considerable_amount : int
        The threshold amount that, when the negative of investment's net difference
        (earnings minus cost) is equal or greater, identifies the investment as just paid.

    Returns
    -------
    list[str]
        List of 'RemateID's identified as just paid based on the criteria.

    Examples
    --------
    >>> data = {'RemateID': ['ID1', 'ID2'],
                'Abono': [1000, 50],
                'Cargo': [1500, 40]}
    >>> df = pd.DataFrame(data)
    >>> extract_just_payed(df, ['ID1', 'ID2'], 400)
    # Returns ['ID1'] since its net investment difference is more than 400 in the negative.
    """
    df_not_in_flows = df[df["RemateID"].isin(not_present_in_flows_ids)]

    dfg = df_not_in_flows.groupby("RemateID")

    just_payed = set()
    for group_key, df_group in dfg:
        earnings = df_group["Abono"].sum()
        cost = df_group["Cargo"].sum()
        investment_diff = earnings - cost

        if investment_diff <= -1 * abs(considerable_amount):
            just_payed.add(group_key)

    return list(just_payed)


def extract_uncollectibles(df: pd.DataFrame, grace_period_days: int) -> list[str]:
    """
    Extracts IDs of investments considered uncollectible based on earnings, costs, and grace period.

    This function groups a DataFrame by 'RemateID' and determines if each grouped investment
    is uncollectible. An investment is considered uncollectible if the difference between
    its total earnings ('Abono') and total costs ('Cargo') is negative beyond a certain threshold,
    and the latest date ('Fecha') of the group's records is past the specified grace period.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing investment data, with columns 'RemateID', 'Abono', 'Cargo',
        and 'Fecha'.

    grace_period_days : int
        The number of days defining the grace period. Investments with their latest date beyond
        this period are considered for being marked as uncollectible.

    Returns
    -------
    list[str]
        A list of 'RemateID's that are considered uncollectible based on the specified criteria.

    Examples
    --------
    >>> data = {'RemateID': ['ID1', 'ID2'],
                'Abono': [500, 50],
                'Cargo': [1500, 100],
                'Fecha': [pd.Timestamp('2023-01-01'), pd.Timestamp('2023-06-01')]}
    >>> df = pd.DataFrame(data)
    >>> extract_uncollectibles(df, 30)
    # Returns a list of IDs which are considered uncollectible.
    """
    uncollectible_ids = set()

    dfg = df.groupby("RemateID")
    for group_key, df_group in dfg:
        earnings = df_group["Abono"].sum()
        cost = df_group["Cargo"].sum()
        date = df_group["Fecha"].max()

        investment_diff = earnings - cost

        if investment_diff <= -1 * abs(1000) and some_utils.is_date_past_grace_period(
            grace_period_days, date
        ):
            uncollectible_ids.add(group_key)

    # negative_investments

    return list(uncollectible_ids)
